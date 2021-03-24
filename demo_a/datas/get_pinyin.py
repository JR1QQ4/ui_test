#!/usr/bin/python
# -*- coding:utf-8 -*-
from pprint import pprint
import random

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from xpinyin import Pinyin
import openpyxl

# words_list = []
# pinyin_list = []
# p = Pinyin()
# for word in words_list:
#     pprint(p.get_pinyin(word).replace("-", ""))

with open("a.txt", "r", encoding="utf-8") as f:
    arr = f.read().split("\n")
    arr1 = []
    p = Pinyin()
    for word in arr:
        arr1.append(p.get_pinyin(word).replace("-", ""))

excel_path = r"C:\Users\东夋壬\Desktop\通讯录批量导入模板.xlsx"
wb: Workbook = openpyxl.load_workbook(excel_path)
sh: Worksheet = wb['成员列表']

name_cols = 1
count_cols = 2
part_cols = 5
part_arr = [
    "奇迹再现/运营部/产品运营",
    "奇迹再现/运营部/内容运营",
    "奇迹再现/运营部/活动策划",
    "奇迹再现/运营部/新媒体",
    "奇迹再现/研发部/前端",
    "奇迹再现/研发部/后端",
    "奇迹再现/研发部/测试",
    "奇迹再现/研发部/运维",
    "奇迹再现/产品部/UI设计师",
    "奇迹再现/产品部/视觉设计师",
    "奇迹再现/市场部/渠道",
    "奇迹再现/市场部/商务",
    "奇迹再现/市场部/推广",
    "奇迹再现/综合部/财务部",
    "奇迹再现/综合部/人事部",
    "奇迹再现/综合部/行政部"
]
sex_cols = 6
phone_cols = 7
start_row = 130
start_phone = 13800000120
for i in range(len(arr)):
    # sh.cell(row=130, column=6).value = arr1[0]
    sh.cell(row=start_row, column=name_cols).value = arr[i]
    sh.cell(row=start_row, column=count_cols).value = arr1[i]
    sh.cell(row=start_row, column=sex_cols).value = random.choice(["男", "女"])
    sh.cell(row=start_row, column=part_cols).value = random.choice(part_arr)
    sh.cell(row=start_row, column=phone_cols).value = start_phone
    start_row += 1
    start_phone += 1
wb.save(excel_path)
